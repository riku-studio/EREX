#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PSTファイルおよびMSGファイルからメールを抽出し、Excelファイルに出力するツール
"""

import os
import sys
import platform
from datetime import datetime
from pathlib import Path
import argparse
import glob

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("エラー: 必要なライブラリがインストールされていません。")
    print("インストール方法: pip install pandas openpyxl")
    sys.exit(1)

try:
    import extract_msg
except ImportError:
    print("エラー: extract_msgライブラリがインストールされていません。")
    print("インストール方法: pip install extract-msg")
    sys.exit(1)

# Windows環境でwin32comをインポート（PSTファイル処理用）
IS_WINDOWS = platform.system() == 'Windows'
if IS_WINDOWS:
    try:
        import win32com.client
        import pythoncom
        WIN32COM_AVAILABLE = True
    except ImportError:
        WIN32COM_AVAILABLE = False
        print("警告: win32comがインストールされていません。PSTファイルの処理には必要です。")
        print("インストール方法: pip install pywin32")
else:
    WIN32COM_AVAILABLE = False


class MSGExtractor:
    """MSGファイルからメールを抽出するクラス"""
    
    def __init__(self, msg_file_path):
        """
        初期化
        
        Args:
            msg_file_path: MSGファイルのパス
        """
        if not os.path.exists(msg_file_path):
            raise FileNotFoundError(f"MSGファイルが見つかりません: {msg_file_path}")
        
        self.msg_file_path = msg_file_path
    
    def extract_message_data(self, folder_name=''):
        """
        MSGファイルからメールデータを抽出
        
        Args:
            folder_name: フォルダ名（オプション）
        
        Returns:
            メッセージデータの辞書
        """
        try:
            msg = extract_msg.Message(self.msg_file_path)
            
            # 件名
            subject = msg.subject or ''
            
            # 送信者情報
            sender_name = msg.sender or ''
            sender_email = msg.senderEmail or ''
            
            # 受信者情報
            recipients = []
            if msg.to:
                recipients.extend([r.strip() for r in msg.to.split(';') if r.strip()])
            if msg.cc:
                recipients.extend([f"CC: {r.strip()}" for r in msg.cc.split(';') if r.strip()])
            if msg.bcc:
                recipients.extend([f"BCC: {r.strip()}" for r in msg.bcc.split(';') if r.strip()])
            
            # 日時
            delivery_time = ''
            if msg.date:
                try:
                    if isinstance(msg.date, datetime):
                        delivery_time = msg.date.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        delivery_time = str(msg.date)
                except:
                    delivery_time = str(msg.date) if msg.date else ''
            
            creation_time = ''
            if hasattr(msg, 'creationTime') and msg.creationTime:
                try:
                    if isinstance(msg.creationTime, datetime):
                        creation_time = msg.creationTime.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        creation_time = str(msg.creationTime)
                except:
                    creation_time = str(msg.creationTime) if msg.creationTime else ''
            
            # 本文
            body = ''
            if msg.body:
                body = msg.body
            elif hasattr(msg, 'htmlBody') and msg.htmlBody:
                # HTMLからテキストを抽出（簡易版）
                import re
                body = re.sub(r'<[^>]+>', '', msg.htmlBody)
            
            # 優先度
            priority_str = '通常'
            if hasattr(msg, 'priority'):
                priority_map = {0: '通常', 1: '高', 2: '高', -1: '低'}
                priority_str = priority_map.get(msg.priority, '通常')
            
            # 既読/未読
            is_read = '未読'
            if hasattr(msg, 'read') and msg.read:
                is_read = '既読'
            
            # フラグ
            is_flagged = ''
            if hasattr(msg, 'flag') and msg.flag:
                is_flagged = 'フラグ付き'
            
            msg.close()
            
            return {
                'フォルダ': folder_name or os.path.dirname(self.msg_file_path),
                '件名': subject,
                '送信者名': sender_name,
                '送信者メールアドレス': sender_email,
                '受信者': '; '.join(recipients),
                '配信日時': delivery_time,
                '作成日時': creation_time,
                '本文': body[:5000] if body else '',  # 最初の5000文字
                '優先度': priority_str,
                '既読': is_read,
                'フラグ': is_flagged,
            }
        except Exception as e:
            print(f"警告: MSGファイルの抽出中にエラーが発生しました ({self.msg_file_path}): {e}")
            return None


class PSTExtractor:
    """PSTファイルからメールを抽出するクラス（Windows環境のみ）"""
    
    def __init__(self, pst_file_path):
        """
        初期化
        
        Args:
            pst_file_path: PSTファイルのパス
        """
        if not os.path.exists(pst_file_path):
            raise FileNotFoundError(f"PSTファイルが見つかりません: {pst_file_path}")
        
        if not IS_WINDOWS or not WIN32COM_AVAILABLE:
            raise RuntimeError(
                "PSTファイルの処理にはWindows環境とpywin32が必要です。\n"
                "macOS/Linux環境では、PSTファイルをMSGファイルにエクスポートしてから処理してください。"
            )
        
        self.pst_file_path = os.path.abspath(pst_file_path)
        self.messages = []
    
    def extract_messages(self, folder=None):
        """
        PSTファイルからメールを抽出
        
        Args:
            folder: 抽出するフォルダ（Noneの場合はすべて）
        
        Returns:
            メールリスト
        """
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            
            # PSTファイルを開く
            try:
                pst = namespace.OpenSharedItem(self.pst_file_path)
            except:
                # PSTファイルをデータストアとして追加
                namespace.AddStore(self.pst_file_path)
                # 開いたストアを取得
                stores = namespace.Stores
                pst_store = None
                for store in stores:
                    if store.FilePath == self.pst_file_path:
                        pst_store = store
                        break
                
                if not pst_store:
                    raise RuntimeError(f"PSTファイルを開けませんでした: {self.pst_file_path}")
                
                root_folder = pst_store.GetRootFolder()
                self._extract_from_folder(root_folder, folder)
                namespace.RemoveStore(root_folder)
                return self.messages
            
            # 単一のアイテムとして開かれた場合
            if hasattr(pst, 'Class'):
                if pst.Class == 43:  # olMail
                    msg_data = self._extract_message_from_item(pst, 'Root')
                    if msg_data:
                        self.messages.append(msg_data)
                elif pst.Class == 2:  # olFolder
                    self._extract_from_folder(pst, folder)
            else:
                self._extract_from_folder(pst, folder)
            
        finally:
            pythoncom.CoUninitialize()
        
        return self.messages
    
    def _extract_from_folder(self, folder, target_folder_name=None):
        """
        フォルダからメールを再帰的に抽出
        
        Args:
            folder: フォルダオブジェクト
            target_folder_name: 対象フォルダ名（Noneの場合はすべて）
        """
        try:
            folder_name = folder.Name
        except:
            folder_name = 'Root'
        
        # 特定のフォルダを指定している場合
        if target_folder_name and folder_name != target_folder_name:
            # サブフォルダをチェック
            try:
                for subfolder in folder.Folders:
                    self._extract_from_folder(subfolder, target_folder_name)
            except:
                pass
            return
        
        # メッセージを取得
        try:
            items = folder.Items
            for item in items:
                try:
                    if item.Class == 43:  # olMail
                        msg_data = self._extract_message_from_item(item, folder_name)
                        if msg_data:
                            self.messages.append(msg_data)
                except Exception as e:
                    # 個々のアイテムの処理エラーは無視
                    pass
        except Exception as e:
            print(f"警告: フォルダ '{folder_name}' の処理中にエラー: {e}")
        
        # サブフォルダを再帰的に処理
        try:
            for subfolder in folder.Folders:
                self._extract_from_folder(subfolder, target_folder_name)
        except:
            pass
    
    def _extract_message_from_item(self, item, folder_name):
        """
        メッセージアイテムからデータを抽出
        
        Args:
            item: メッセージアイテム
            folder_name: フォルダ名
        
        Returns:
            メッセージデータの辞書
        """
        try:
            # 件名
            subject = getattr(item, 'Subject', '') or ''
            
            # 送信者情報
            sender_name = ''
            sender_email = ''
            try:
                sender = item.SenderName
                sender_name = sender if sender else ''
            except:
                pass
            
            try:
                sender_email = item.SenderEmailAddress if hasattr(item, 'SenderEmailAddress') else ''
            except:
                pass
            
            # 受信者情報
            recipients = []
            try:
                if hasattr(item, 'To') and item.To:
                    recipients.append(item.To)
            except:
                pass
            
            try:
                if hasattr(item, 'CC') and item.CC:
                    recipients.append(f"CC: {item.CC}")
            except:
                pass
            
            try:
                if hasattr(item, 'BCC') and item.BCC:
                    recipients.append(f"BCC: {item.BCC}")
            except:
                pass
            
            # 日時
            delivery_time = ''
            try:
                if hasattr(item, 'ReceivedTime') and item.ReceivedTime:
                    dt = item.ReceivedTime
                    if isinstance(dt, datetime):
                        delivery_time = dt.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        delivery_time = str(dt)
            except:
                pass
            
            creation_time = ''
            try:
                if hasattr(item, 'CreationTime') and item.CreationTime:
                    dt = item.CreationTime
                    if isinstance(dt, datetime):
                        creation_time = dt.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        creation_time = str(dt)
            except:
                pass
            
            # 本文
            body = ''
            try:
                body = item.Body if hasattr(item, 'Body') else ''
            except:
                pass
            
            # 優先度
            priority_str = '通常'
            try:
                if hasattr(item, 'Importance'):
                    priority_map = {0: '低', 1: '通常', 2: '高'}
                    priority_str = priority_map.get(item.Importance, '通常')
            except:
                pass
            
            # 既読/未読
            is_read = '未読'
            try:
                if hasattr(item, 'UnRead'):
                    is_read = '既読' if not item.UnRead else '未読'
            except:
                pass
            
            # フラグ
            is_flagged = ''
            try:
                if hasattr(item, 'FlagStatus') and item.FlagStatus and item.FlagStatus != 0:
                    is_flagged = 'フラグ付き'
            except:
                pass
            
            return {
                'フォルダ': folder_name,
                '件名': subject,
                '送信者名': sender_name,
                '送信者メールアドレス': sender_email,
                '受信者': '; '.join(recipients),
                '配信日時': delivery_time,
                '作成日時': creation_time,
                '本文': body[:5000] if body else '',  # 最初の5000文字
                '優先度': priority_str,
                '既読': is_read,
                'フラグ': is_flagged,
            }
        except Exception as e:
            print(f"警告: メッセージの抽出中にエラーが発生しました: {e}")
            return None


def extract_from_file(file_path, folder_name=None):
    """
    ファイルからメールを抽出（PSTまたはMSG）
    
    Args:
        file_path: ファイルパス
        folder_name: フォルダ名（PSTファイルの場合のみ有効）
    
    Returns:
        メッセージリスト
    """
    file_path = os.path.abspath(file_path)
    file_ext = os.path.splitext(file_path)[1].lower()
    
    messages = []
    
    if file_ext == '.msg':
        # MSGファイル
        extractor = MSGExtractor(file_path)
        msg_data = extractor.extract_message_data()
        if msg_data:
            messages.append(msg_data)
    elif file_ext == '.pst':
        # PSTファイル
        if not IS_WINDOWS or not WIN32COM_AVAILABLE:
            raise RuntimeError(
                "PSTファイルの処理にはWindows環境とpywin32が必要です。\n"
                "macOS/Linux環境では、PSTファイルをMSGファイルにエクスポートしてから処理してください。"
            )
        extractor = PSTExtractor(file_path)
        messages = extractor.extract_messages(folder=folder_name)
    else:
        raise ValueError(f"サポートされていないファイル形式です: {file_ext}")
    
    return messages


def extract_from_directory(directory_path):
    """
    ディレクトリ内のMSGファイルからメールを抽出
    
    Args:
        directory_path: ディレクトリパス
    
    Returns:
        メッセージリスト
    """
    messages = []
    directory_path = os.path.abspath(directory_path)
    
    # MSGファイルを検索
    msg_files = glob.glob(os.path.join(directory_path, '**', '*.msg'), recursive=True)
    msg_files.extend(glob.glob(os.path.join(directory_path, '**', '*.MSG'), recursive=True))
    
    print(f"見つかったMSGファイル数: {len(msg_files)}")
    
    for msg_file in msg_files:
        try:
            extractor = MSGExtractor(msg_file)
            msg_data = extractor.extract_message_data()
            if msg_data:
                messages.append(msg_data)
        except Exception as e:
            print(f"警告: {msg_file} の処理中にエラー: {e}")
            continue
    
    return messages


def export_to_excel(messages, output_file):
    """
    メッセージをExcelファイルに出力
    
    Args:
        messages: メッセージリスト
        output_file: 出力ファイルパス
    """
    if not messages:
        print("警告: 出力するメッセージがありません。")
        return
    
    # DataFrameを作成
    df = pd.DataFrame(messages)
    
    # Excelファイルに出力
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='メール一覧', index=False)
        
        # ワークシートを取得してスタイルを適用
        worksheet = writer.sheets['メール一覧']
        
        # ヘッダーのスタイル
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 列幅の調整
        column_widths = {
            'A': 20,  # フォルダ
            'B': 40,  # 件名
            'C': 20,  # 送信者名
            'D': 30,  # 送信者メールアドレス
            'E': 40,  # 受信者
            'F': 20,  # 配信日時
            'G': 20,  # 作成日時
            'H': 60,  # 本文
            'I': 10,  # 優先度
            'J': 10,  # 既読
            'K': 10,  # フラグ
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 本文列の折り返し
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=8, max_col=8):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 他の列のスタイル
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column != 8:  # 本文以外
                    cell.alignment = Alignment(vertical='top')
    
    print(f"Excelファイルを出力しました: {output_file}")
    print(f"抽出されたメール数: {len(messages)}")


def main():
    """メイン処理"""
    parser = argparse.ArgumentParser(
        description='PSTファイルおよびMSGファイルからメールを抽出し、Excelファイルに出力するツール'
    )
    parser.add_argument('input_path', help='入力PSTファイル、MSGファイル、またはディレクトリのパス')
    parser.add_argument('-o', '--output', help='出力Excelファイルのパス（省略時は自動生成）')
    parser.add_argument('-f', '--folder', help='抽出するフォルダ名（PSTファイルの場合のみ有効）')
    
    args = parser.parse_args()
    
    input_path = os.path.abspath(args.input_path)
    
    if not os.path.exists(input_path):
        print(f"エラー: 入力パスが存在しません: {input_path}")
        sys.exit(1)
    
    # 出力ファイル名を決定
    if args.output:
        output_file = args.output
    else:
        if os.path.isfile(input_path):
            base_name = Path(input_path).stem
        else:
            base_name = Path(input_path).name
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"{base_name}_extracted_{timestamp}.xlsx"
    
    try:
        messages = []
        
        if os.path.isfile(input_path):
            # ファイルの場合
            file_ext = os.path.splitext(input_path)[1].lower()
            print(f"ファイルを読み込み中: {input_path}")
            
            if file_ext == '.msg':
                messages = extract_from_file(input_path)
            elif file_ext == '.pst':
                if not IS_WINDOWS or not WIN32COM_AVAILABLE:
                    print("エラー: PSTファイルの処理にはWindows環境とpywin32が必要です。")
                    print("macOS/Linux環境では、PSTファイルをMSGファイルにエクスポートしてから処理してください。")
                    sys.exit(1)
                messages = extract_from_file(input_path, folder_name=args.folder)
            else:
                print(f"エラー: サポートされていないファイル形式です: {file_ext}")
                sys.exit(1)
        
        elif os.path.isdir(input_path):
            # ディレクトリの場合（MSGファイルを検索）
            print(f"ディレクトリからMSGファイルを検索中: {input_path}")
            messages = extract_from_directory(input_path)
        else:
            print(f"エラー: 無効な入力パスです: {input_path}")
            sys.exit(1)
        
        if not messages:
            print("警告: 抽出されたメッセージがありません。")
            return
        
        print(f"抽出されたメッセージ数: {len(messages)}")
        
        # Excelに出力
        export_to_excel(messages, output_file)
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()